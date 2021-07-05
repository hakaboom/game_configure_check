import openpyxl
import os
from openpyxl.utils import get_column_letter


class Pyxl(object):
    def __init__(self, file_name, path='', init: bool=True):
        """
        初始化实例
        :param file_name: 表格名称
        :param path: 表格路径
        :param init: False将会读取已有的表, True将会创建新的表s
        """
        self.xls_name = file_name
        self.path = path
        self.wb = self.init_workbook(init)
        self.sheet = self.wb.active
        self.ignore_lines = 5  # 获取数据时,跳过不会读取的行
        self.end_tag_index = self.sheet.max_row
        self.set_end_tag_index()

    def init_workbook(self, flags):
        path = os.path.join(self.path, self.xls_name)
        try:
            open(path, 'r')
        except FileNotFoundError:
            return openpyxl.Workbook()
        else:
            if flags:
                return openpyxl.Workbook()
            else:
                return openpyxl.load_workbook(path)

    def get_row_value_list(self, rowx: int, start_colx:int = 0, end_cols: int = None):
        """
        获取表格内某一行的数据
        :param rowx: 在表格中的行数
        :param start_colx: 左切片
        :param end_colx: 右切片
        :return: 这一行的全部数据
        """
        return [v.value for v in self.sheet[rowx]][start_colx:end_cols]

    def get_col_value_list(self, colx: int, start_rowx: int = 0, end_rowx: int = None) -> list:
        """
        获取表格内某一列的数据
        :param colx: 在表格中的列数
        :param start_rowx: 左切片
        :param end_rowx: 右切片
        :return: 这一列的全部数据
        """
        return [v.value for v in self.sheet[get_column_letter(colx)]][start_rowx:end_rowx]

    def get_cell(self, row_index: int, column_index: int):
        """
        获取某一行某一列的单元格的索引
        :param row_index: 在表格中的行数
        :param column_index: 在表格中的列数
        :return: 这一个单元格内的数据
        """
        return self.sheet.cell(row_index, column_index)

    def get_cell_value(self, row_index: int, column_index: int):
        """
        获取某一行某一列的单元格内数据
        :param row_index: 在表格中的行数
        :param column_index: 在表格中的列数
        :return: 这一个单元格内的数据
        """
        return self.get_cell(row_index, column_index).value

    def get_col_list_by_name(self, name: str) -> list:
        head_list = self.get_row_value_list(self.ignore_lines)
        index = head_list.index(name) + 1if name in head_list else None
        if index:
            return self.get_col_value_list(index)[self.ignore_lines:]
        else:
            raise ValueError('{xlsName}没有{name}列'.format(xlsName=self.xls_name, name=name))

    def get_col_number_by_name(self, name: str) -> int:
        head_list = self.get_row_value_list(self.ignore_lines)
        index = head_list.index(name) + 1 if name in head_list else None
        if index:
            return index
        else:
            raise ValueError('{xlsName}没有{name}列'.format(xlsName=self.xls_name, name=name))

    def write_value_in_cell(self, value, row_index: int, column_index: int):
        """
        写入数据到指定的单元格上
        :param value: 需要填入的数据
        :param row_index:  在表格中的行数
        :param column_index: 在表格中的列数
        :return: None
        """
        cell = self.get_cell(row_index, column_index)
        cell.value = value

    def write_value_to_col(self, value, colx: int = None, start_row: int = None):
        """
        写入数据到指定列上
        :param value: 需要填入的数据
        :param colx: 在表格中的列数
        :param start_row: 初始的索引
        :return: Nones
        """
        if not isinstance(value, (list, tuple)):
            raise ValueError('write_value_to_col error, 传入错误的值', value)

        start_row = start_row or 1
        for v in value:
            cell = self.get_cell(column_index=colx, row_index=start_row)
            cell.value = v
            start_row += 1

    def write_value_to_row(self, value, rowx: int = None, start_col: int = None, style=None):
        """
        写入数据到指定行上
        :param value: 需要填入的数据
        :param rowx: 在表格中的行数
        :param start_col: 初始的索引
        :return: Nones
        """
        if not isinstance(value, (list, tuple)):
            raise ValueError('write_value_to_col error, 传入错误的值', value)

        start_col = start_col or 1
        start_row = rowx or self.sheet.max_row + 1
        for v in value:
            cell = self.get_cell(column_index=start_col, row_index=start_row)
            cell.value = v
            if style:
                cell.style = style
            start_col += 1

    def set_end_tag_index(self):
        self.end_tag_index = self.get_end_tag_index()

    def get_end_tag_index(self):
        """ 数据表中第一列的最后一行都有 #END_TAG#, 获取时要舍弃这一行的数据 """
        try:
            firest_col = self.get_col_value_list(1, end_rowx=self.sheet.max_row)
            end_tag_index = firest_col.index('#END_TAG#')
        except ValueError:
            return self.sheet.max_row
        else:
            return end_tag_index

    def save(self):
        self.wb.save(os.path.join(self.path, self.xls_name))